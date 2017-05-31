from openpyxl import load_workbook, worksheet
from Student import Student

EMAIL_COL = 4
FIRSTNAME_COL = 2
LASTNAME_COL = 3
SCORE_COL = 15
COURSE_COL = 12

def WriteData(filename, students):
    Update_Existing_Students(filename, students)

def Update_Existing_Students(filename, students):

    wb = load_workbook(filename=filename)
    sheet = wb.get_sheet_by_name('Attendees')

    for idx, row in enumerate(sheet.iter_rows(), start=2):
        email_cell = sheet.cell(row=idx, column=EMAIL_COL).value

        if not email_cell:
            continue  # email is empty move on

        # grab score and course from spreadsheet
        score_cell = sheet.cell(row=idx, column=SCORE_COL)
        course_cell = sheet.cell(row=idx, column=COURSE_COL)

        # query student data loaded from .org sheets and look for a match
        # ! MUST CHECK FOR MULTIPLES -- MAYBE ADD COURSE !
        student = Student.find_by_email(email_cell)

        if student:
            # update the student scores
            for s in student:
                if s.score and not score_cell.value:
                    score_cell.value = s.score
                    print('Score Update: {} {} {} {}'.format(s.firstname, s.lastname, s.email, s.score))

    wb.save(filename)

def AddNewStudents(filename, students):

    for student in students:
        print('students')

    wb = load_workbook(filename=filename)
    sheet = wb.get_sheet_by_name('Attendees')

    for idx, row in enumerate(sheet.iter_rows(), start=2):
        email_cell = sheet.cell(row=idx, column=EMAIL_COL).value

        if not email_cell:
            continue  # email is empty move on

        # grab score and course from spreadsheet
        score_cell = sheet.cell(row=idx, column=SCORE_COL)
        course_cell = sheet.cell(row=idx, column=COURSE_COL)

        # query student data loaded from .org sheets and look for a match
        # ! MUST CHECK FOR MULTIPLES -- MAYBE ADD COURSE !
        student = Student.find_by_email(email_cell)

        if student:
            # update the student scores
            print('here')
            for s in student:
                if s.score and not score_cell.value:
                    score_cell.value = s.score
                    print('New Score: {} {}'.format(s.email, s.score))
        else:
            print('here 2')
            # add student record
            firstname_cell = sheet.cell(row=idx, column=FIRSTNAME_COL)
            lastname_cell = sheet.cell(row=idx, column=LASTNAME_COL)
            sheet.append(student)
            # print('Added student: {} {}'.format(s.email, s.score))

    wb.save(filename)